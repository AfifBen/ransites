import io
import re
import csv
import math
import json
import threading
import uuid
from datetime import datetime
from urllib.parse import urlencode
from urllib.request import urlopen

from flask import Blueprint, current_app, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload

from app import db
from app.models import Region, Wilaya, Commune, Site, Antenna, Supplier, Sector, Mapping, Cell, Cell2G, Cell3G, Cell4G
from app.security import admin_required, append_audit_event, login_required, csrf_protect, get_accessible_site_ids
from app.ran_reference import build_ran_reference_map

main_bp = Blueprint('main', __name__)

_cell_sector_sync_jobs = {}
_cell_sector_sync_lock = threading.Lock()
_site_altitude_sync_jobs = {}
_site_altitude_sync_lock = threading.Lock()


def _haversine_km(lat1, lon1, lat2, lon2):
    # Great-circle distance between two GPS points (in km).
    r = 6371.0
    p1 = math.radians(float(lat1))
    p2 = math.radians(float(lat2))
    dlat = math.radians(float(lat2) - float(lat1))
    dlon = math.radians(float(lon2) - float(lon1))
    a = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return r * c


def _fetch_ground_altitude(latitude, longitude):
    # Best-effort elevation lookup from Open-Elevation API.
    try:
        query = urlencode({"locations": f"{latitude},{longitude}"})
        url = f"https://api.open-elevation.com/api/v1/lookup?{query}"
        with urlopen(url, timeout=3) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        results = payload.get("results") or []
        if not results:
            return None
        val = results[0].get("elevation")
        return float(val) if val is not None else None
    except Exception:
        return None

IMPORT_TEMPLATE_SPECS = {
    "sites": {
        "sheets": {
            "sites": [
                "site_code",
                "site_name",
                "commune_id",
                "supplier_name",
                "latitude",
                "longitude",
                "laltitude",
                "addresses",
                "altitude",
                "support_nature",
                "support_type",
                "support_hight",
                "Comments",
            ]
        }
    },
    "sectors": {
        "sheets": {
            "sectors": [
                "code_sector",
                "azimuth",
                "hba",
                "code_site",
                "coverage_goal",
                "comments",
            ]
        }
    },
    "cells": {
        "sheets": {
            "2G": [
                "CELLNAME",
                "TECHNOLOGY",
                "FREQUENCY",
                "ANTENNA_TECH",
                "MECHANICALTILT",
                "ELECTRICALTILT",
                "ANTENNA",
                "BSC",
                "LAC",
                "RAC",
                "BSIC",
                "BCCH",
                "CI",
            ],
            "3G": [
                "CELLNAME",
                "TECHNOLOGY",
                "FREQUENCY",
                "ANTENNA_TECH",
                "MECHANICALTILT",
                "ELECTRICALTILT",
                "ANTENNA",
                "RNC",
                "LAC",
                "RAC",
                "PSC",
                "DLARFCN",
                "CI",
            ],
            "4G": [
                "CELLNAME",
                "TECHNOLOGY",
                "FREQUENCY",
                "ANTENNA_TECH",
                "MECHANICALTILT",
                "ELECTRICALTILT",
                "ANTENNA",
                "ENODEB",
                "TAC",
                "RSI",
                "PCI",
                "EARFCN",
                "CI",
            ],
            "5G": [
                "CELLNAME",
                "TECHNOLOGY",
                "FREQUENCY",
                "ANTENNA_TECH",
                "MECHANICALTILT",
                "ELECTRICALTILT",
                "ANTENNA",
                "GNODEB",
                "LAC",
                "RSI",
                "PCI",
                "ARFCN",
                "CI",
            ],
        }
    },
    "mapping": {
        "sheets": {
            "mapping": [
                "MAP_ID",
                "CELL_CODE",
                "ANTENNA_TECH",
                "BAND",
                "SECTOR_CODE",
                "TECHNOLOGY",
            ]
        }
    },
    "antennas": {
        "sheets": {
            "antennas": [
                "Supplier",
                "Model",
                "Frequency",
                "HBEAMWIDTH",
                "VBEAMWIDTH",
                "Name",
                "Port",
                "Type",
                "GAIN",
            ]
        }
    },
    "suppliers": {"sheets": {"suppliers": ["SUPPLIER_NAME"]}},
    "regions": {"sheets": {"regions": ["name"]}},
    "wilayas": {"sheets": {"wilayas": ["wilaya_code", "wilaya_name", "region_name"]}},
    "communes": {"sheets": {"communes": ["commune_id", "commune_name", "wilaya_name"]}},
}

IMPORT_TEMPLATE_EXAMPLES = {
    "sites": {
        "sites": [
            "C28SU217",
            "SUT_SMAYER_ZRAYEF",
            2801,
            "Nokia",
            35.82046,
            4.78125,
            35.82046,
            "MSILA",
            520,
            "Wall Mounted",
            "Directional",
            18,
            "Example site",
        ]
    },
    "sectors": {
        "sectors": [
            "C28SU217_1",
            120,
            20,
            "C28SU217",
            "Coverage",
            "Example sector",
        ]
    },
    "cells": {
        "2G": ["2C28SU217_1", "2G", "GSM900", "2G", 2, 1, "ANT-900-65", "BSC01", "1201", "1", "63", 62, 2801001],
        "3G": ["3C28SU217_1", "3G", "U2100", "3G", 2, 1, "ANT-2100-65", "RNC01", "1201", "1", 245, "10612", 2801001],
        "4G": ["4C28SU217_1", "4G", "L1800", "4G", 2, 1, "ANT-L18-65", "ENB2801", "2801", "12", 321, "1650", 2801001],
        "5G": ["5C28SU217_1", "5G", "N78", "5G", 2, 1, "ANT-N78-64T", "GNB2801", "1201", "15", 501, "636666", 2801001],
    },
    "mapping": {"mapping": ["MAP_0001", "1", "4G", "L1800", "1", "4G"]},
    "antennas": {"antennas": ["Nokia", "ANT-L18-65", 1800, 65, 6.5, "Panel 18", 2, "Panel", 17.8]},
    "suppliers": {"suppliers": ["Nokia"]},
    "regions": {"regions": ["Hauts Plateaux"]},
    "wilayas": {"wilayas": [28, "MSILA", "Hauts Plateaux"]},
    "communes": {"communes": [2801, "MSILA", "MSILA"]},
}

ALLPLAN_HEADERS = {
    "2G": [
        "CELLNAME", "LONGITUDE", "LATITUDE", "AZIMUTH", "HEIGHT", "ANTENNA",
        "ELECTRICALTILT", "MECHANICALTILT", "HBEAMWIDTH", "VBEAMWIDTH", "GAIN",
        "SITENAME", "SITE_TYPE", "INDOORFLAG", "SECTORID", "ARFCN", "LAC",
        "CELLIDENTITY", "BCCH", "BSIC(octal)", "Core or Buffer", "New Cell", "COMMUNE",
    ],
    "3G": [
        "CELLNAME", "LONGITUDE", "LATITUDE", "AZIMUTH", "HEIGHT", "ANTENNA",
        "ELECTRICALTILT", "MECHANICALTILT", "HBEAMWIDTH", "VBEAMWIDTH", "GAIN",
        "SITENAME", "SITE_TYPE", "INDOORFLAG", "SECTORID", "DL UARFCN", "LAC",
        "RNCID", "UTRANCELLIDENTITY", "Core or Buffer", "New Cell", "COMMUNE",
    ],
    "4G": [
        "CELLNAME", "LONGITUDE", "LATITUDE", "AZIMUTH", "HEIGHT", "ANTENNA",
        "ELECTRICALTILT", "MECHANICALTILT", "HBEAMWIDTH", "VBEAMWIDTH", "ANTENNAGAIN",
        "ENODEBNAME", "ENODEBID", "CI", "PCI", "DL EARFCN", "SITENAME", "SITE_TYPE",
        "INDOORFLAG", "SECTORID", "Core or Buffer", "New Cell", "COMMUNE",
    ],
}

LBS_HEADERS = {
    "2G": [
        "CID", "Site Name", "CellName", "Longitude", "Latitude", "AntennaType", "MaxCellRadius",
        "AntennaGain", "BSNominalPower", "BSPowerBCCH", "TAlimit", "AntennaSpec", "HeightAGL",
        "DownTilt", "MechanicalTilt", "ElectricalTilt", "Azimuth", "HorizBeamWidth", "LAC",
        "NCC", "BCC", "BCCH", "MCC", "MNC", "FrequencyBand", "Node Name",
    ],
    "3G": [
        "CID", "SiteName", "CellName", "Longitude", "Latitude", "AntennaType", "MaxCellRadius",
        "Gain", "BSNominalPower", "AntennaSpec", "HeightAGL", "DownTilt", "Azimuth",
        "MechanicalTilt", "ElectricalTilt", "HorizBeamWidth", "LAC", "SAC", "RNCid",
        "PSC", "UARFCN", "FrequencyBand", "NodeName",
    ],
    "4G": [
        "Site Name", "CellName", "Longitude", "Latitude", "AntennaType", "MaxCellRadius",
        "AntennaGain", "BSNominalPower", "AntennaSpec", "HeightAGL", "DownTilt",
        "MechanicalTilt", "ElectricalTilt", "Azimuth", "HorizBeamWidth", "TAC ",
        "E-UTRANCELLID", "EARFCN", "FrequencyBand", "PCI", "NodeName",
    ],
}

def get_stats():
    return {
        'total_regions': Region.query.count(),
        'total_wilayas': Wilaya.query.count(),
        'total_communes': Commune.query.count(),
        'total_suppliers': Supplier.query.count(),
        'total_antennas': Antenna.query.count(),
        'total_cell_id_mapping': Mapping.query.count(),
        'total_cells': Cell.query.count(),
        'total_sectors': Sector.query.count(),
        'total_sites': Site.query.count(),
    }


def _normalize_tech_for_dashboard(value):
    tech = _normalize_tech(value)
    return tech if tech else "Other"


def _dashboard_scope_badge():
    if not getattr(current_user, "is_authenticated", False):
        return ""
    if getattr(current_user, "is_admin_user", False):
        return "Scoped to: All Network"

    try:
        region_count = len(getattr(current_user, "assigned_regions", []) or [])
    except Exception:
        region_count = 0
    try:
        wilaya_count = len(getattr(current_user, "assigned_wilayas", []) or [])
    except Exception:
        wilaya_count = 0
    try:
        site_count = len(getattr(current_user, "assigned_sites", []) or [])
    except Exception:
        site_count = 0
    return f"Scoped to: Regions {region_count} | Wilayas {wilaya_count} | Sites {site_count}"


def get_dashboard_data():
    global_stats = get_stats()

    accessible_sites = get_accessible_site_ids()

    site_base_query = Site.query
    if accessible_sites is not None:
        if not accessible_sites:
            site_base_query = site_base_query.filter(False)
        else:
            site_base_query = site_base_query.filter(Site.id.in_(list(accessible_sites)))

    scoped_site_ids_subq = site_base_query.with_entities(Site.id).subquery()

    total_sites = site_base_query.count()

    sector_base_query = Sector.query.filter(Sector.site_id.in_(db.session.query(scoped_site_ids_subq.c.id)))
    total_sectors = sector_base_query.count()

    cell_base_query = (
        Cell.query
        .outerjoin(Sector, Cell.sector_id == Sector.id)
        .outerjoin(Site, Sector.site_id == Site.id)
        .filter(Site.id.in_(db.session.query(scoped_site_ids_subq.c.id)))
    )
    total_cells = cell_base_query.count()

    if accessible_sites is None:
        scoped_regions = global_stats["total_regions"]
        scoped_wilayas = global_stats["total_wilayas"]
        scoped_communes = global_stats["total_communes"]
        scoped_suppliers = global_stats["total_suppliers"]
        scoped_antennas = global_stats["total_antennas"]
    else:
        scoped_regions = (
            db.session.query(func.count(func.distinct(Wilaya.region_id)))
            .join(Commune, Commune.wilaya_id == Wilaya.id)
            .join(Site, Site.commune_id == Commune.id)
            .filter(Site.id.in_(list(accessible_sites)))
            .scalar()
            or 0
        )
        scoped_wilayas = (
            db.session.query(func.count(func.distinct(Commune.wilaya_id)))
            .join(Site, Site.commune_id == Commune.id)
            .filter(Site.id.in_(list(accessible_sites)))
            .scalar()
            or 0
        )
        scoped_communes = (
            db.session.query(func.count(func.distinct(Site.commune_id)))
            .filter(Site.id.in_(list(accessible_sites)))
            .scalar()
            or 0
        )
        scoped_suppliers = (
            db.session.query(func.count(func.distinct(Site.supplier_id)))
            .filter(Site.id.in_(list(accessible_sites)), Site.supplier_id.isnot(None))
            .scalar()
            or 0
        )
        scoped_antennas = (
            db.session.query(func.count(func.distinct(Cell.antenna_id)))
            .outerjoin(Sector, Cell.sector_id == Sector.id)
            .outerjoin(Site, Sector.site_id == Site.id)
            .filter(Site.id.in_(list(accessible_sites)), Cell.antenna_id.isnot(None))
            .scalar()
            or 0
        )

    stats = {
        "total_sites": total_sites,
        "total_sectors": total_sectors,
        "total_cells": total_cells,
        "total_regions": scoped_regions,
        "total_wilayas": scoped_wilayas,
        "total_communes": scoped_communes,
        "total_suppliers": scoped_suppliers,
        "total_antennas": scoped_antennas,
        "total_cell_id_mapping": global_stats["total_cell_id_mapping"],
    }

    avg_sectors_per_site = round(total_sectors / total_sites, 2) if total_sites else 0
    avg_cells_per_sector = round(total_cells / total_sectors, 2) if total_sectors else 0

    tech_query = cell_base_query.with_entities(Cell.technology, func.count(Cell.id)).group_by(Cell.technology)
    tech_raw = tech_query.all()
    tech_counts = {"2G": 0, "3G": 0, "4G": 0, "Other": 0}
    for tech, count in tech_raw:
        tech_counts[_normalize_tech_for_dashboard(tech)] += count

    tech_distribution = []
    for label in ("2G", "3G", "4G", "Other"):
        value = tech_counts[label]
        pct = round((value * 100.0 / total_cells), 1) if total_cells else 0
        tech_distribution.append({"label": label, "value": value, "pct": pct})

    top_wilayas_query = (
        Wilaya.query.with_entities(Wilaya.name, func.count(Site.id).label("site_count"))
        .join(Commune, Commune.wilaya_id == Wilaya.id)
        .outerjoin(Site, Site.commune_id == Commune.id)
        .filter(Site.id.in_(db.session.query(scoped_site_ids_subq.c.id)))
        .group_by(Wilaya.id, Wilaya.name)
    )
    top_wilayas_raw = top_wilayas_query.order_by(func.count(Site.id).desc(), Wilaya.name.asc()).limit(8).all()
    max_wilaya = max((row.site_count for row in top_wilayas_raw), default=1)
    top_wilayas = [
        {
            "name": row.name,
            "count": row.site_count,
            "pct": round((row.site_count * 100.0 / max_wilaya), 1) if max_wilaya else 0,
        }
        for row in top_wilayas_raw
    ]

    top_suppliers_query = (
        site_base_query.with_entities(
            func.coalesce(Supplier.name, "Unassigned").label("supplier"),
            func.count(Site.id).label("site_count"),
        )
        .outerjoin(Supplier, Site.supplier_id == Supplier.id)
        .group_by(func.coalesce(Supplier.name, "Unassigned"))
    )
    top_suppliers_raw = top_suppliers_query.order_by(func.count(Site.id).desc()).limit(8).all()
    max_supplier = max((row.site_count for row in top_suppliers_raw), default=1)
    top_suppliers = [
        {
            "name": row.supplier,
            "count": row.site_count,
            "pct": round((row.site_count * 100.0 / max_supplier), 1) if max_supplier else 0,
        }
        for row in top_suppliers_raw
    ]

    sites_without_sectors = (
        site_base_query.outerjoin(Sector, Sector.site_id == Site.id)
        .filter(Sector.id.is_(None))
        .count()
    )
    sectors_without_cells = (
        sector_base_query.outerjoin(Cell, Cell.sector_id == Sector.id)
        .filter(Cell.id.is_(None))
        .count()
    )
    # Cells without sector have no site binding; keep count only for global admin view.
    if accessible_sites is None:
        cells_without_sector = Cell.query.filter(Cell.sector_id.is_(None)).count()
    else:
        cells_without_sector = 0

    cells_without_antenna_query = cell_base_query.filter(Cell.antenna_id.is_(None))
    cells_without_antenna = cells_without_antenna_query.count()

    sites_without_supplier = site_base_query.filter(Site.supplier_id.is_(None)).count()

    mapping_codes = {
        (code or "").strip()
        for (code,) in Mapping.query.with_entities(Mapping.cell_code).distinct().all()
        if (code or "").strip()
    }
    mapped_cells = 0
    for (cellname,) in cell_base_query.with_entities(Cell.cellname).all():
        code = _extract_cell_code(cellname)
        if code and code in mapping_codes:
            mapped_cells += 1
    mapping_coverage = round((mapped_cells * 100.0 / total_cells), 1) if total_cells else 0

    data_quality = {
        "sites_without_sectors": sites_without_sectors,
        "sectors_without_cells": sectors_without_cells,
        "cells_without_sector": cells_without_sector,
        "cells_without_antenna": cells_without_antenna,
        "sites_without_supplier": sites_without_supplier,
        "mapped_cells": mapped_cells,
        "mapping_coverage_pct": mapping_coverage,
    }

    return {
        "stats": stats,
        "scope_badge": _dashboard_scope_badge(),
        "avg_sectors_per_site": avg_sectors_per_site,
        "avg_cells_per_sector": avg_cells_per_sector,
        "tech_distribution": tech_distribution,
        "top_wilayas": top_wilayas,
        "top_suppliers": top_suppliers,
        "data_quality": data_quality,
        "charts": {
            "tech_labels": [row["label"] for row in tech_distribution],
            "tech_values": [row["value"] for row in tech_distribution],
            "wilaya_labels": [row["name"] for row in top_wilayas],
            "wilaya_values": [row["count"] for row in top_wilayas],
            "supplier_labels": [row["name"] for row in top_suppliers],
            "supplier_values": [row["count"] for row in top_suppliers],
            "mapping": {
                "mapped": mapped_cells,
                "unmapped": max(total_cells - mapped_cells, 0),
            },
        },
        "generated_at": datetime.utcnow(),
    }


def _normalize_tech(value):
    tech = (value or "").strip().upper()
    if tech.startswith("2") or tech in {"GSM"}:
        return "2G"
    if tech.startswith("3") or tech in {"UMTS", "WCDMA"}:
        return "3G"
    if tech.startswith("4") or tech in {"LTE"}:
        return "4G"
    return None


def _extract_sector_id(cellname):
    if not cellname:
        return None
    parts = str(cellname).rsplit("_", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return int(parts[1])
    return None


def _extract_cell_code(cellname):
    if not cellname:
        return None
    parts = str(cellname).rsplit("_", 1)
    if len(parts) == 2 and parts[1]:
        return parts[1]
    return None


def _parse_cell_list(raw_text):
    tokens = re.split(r"[,\s;]+", raw_text or "")
    ordered = []
    seen = set()
    for token in tokens:
        cell = token.strip()
        if not cell:
            continue
        if cell not in seen:
            seen.add(cell)
            ordered.append(cell)
    return ordered


def _parse_cell_file(file_storage):
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()
    if not raw:
        return []

    values = []
    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            values.append(row[0])
    elif filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            values.append(row[0] if row else None)
    else:
        raise ValueError("Format non supporté. Utilisez .xlsx ou .csv.")

    ordered = []
    seen = set()
    for value in values:
        cell = str(value or "").strip()
        if not cell:
            continue
        if cell.lower() in {"cellname", "cell_name", "cell", "cells"}:
            continue
        if cell not in seen:
            seen.add(cell)
            ordered.append(cell)
    return ordered


def _cell_suffix_int(cellname):
    parts = str(cellname or "").rsplit("_", 1)
    if len(parts) == 2 and str(parts[1]).isdigit():
        return int(parts[1])
    return None


def _freq_band_int(value):
    txt = str(value or "").strip()
    m = re.search(r"(\d{3,5})", txt)
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _antenna_type_for_lbs(site):
    st = (site.support_type or "").lower() if site else ""
    if "indoor" in st:
        return "Indoor"
    return "Macro"


def _ncc_bcc_from_bsic(raw_bsic):
    if raw_bsic is None:
        return None, None


def _lbs_radius_km(tech, band, ran_map):
    try:
        b = int(float(band)) if band is not None else None
    except (TypeError, ValueError):
        b = None
    if b is not None and (tech, b) in ran_map:
        return ran_map[(tech, b)]["radius"]
    if tech == "2G":
        return 35
    if tech == "3G":
        return 25
    if tech == "4G":
        return 20
    return 4


def _lbs_nominal_power_dbm(tech, band, ran_map):
    try:
        b = int(float(band)) if band is not None else None
    except (TypeError, ValueError):
        b = None
    if b is not None and (tech, b) in ran_map:
        return ran_map[(tech, b)]["power"]
    if tech == "2G":
        return 48
    if tech == "3G":
        return 46
    if tech == "4G":
        return 46
    return 46
    try:
        v = int(float(raw_bsic))
        if v < 0:
            return None, None
        return v // 8, v % 8
    except (TypeError, ValueError):
        return None, None


@main_bp.route('/')
@main_bp.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', dashboard=get_dashboard_data())


@main_bp.route('/import_export')
@login_required
@admin_required
def import_export():
    return render_template('import_export.html', title='Import & Export')


@main_bp.route('/site-profile/<int:site_id>', methods=['GET'])
@login_required
def site_profile(site_id):
    # 1) Load site + enforce access scope for non-admin users.
    site = Site.query.get(site_id)
    if not site:
        return jsonify({"success": False, "message": "Site not found."}), 404

    accessible = get_accessible_site_ids()
    if accessible is not None and site.id not in accessible:
        return jsonify({"success": False, "message": "Access denied for this site."}), 403

    # 2) Gather full local topology for the selected site.
    sectors = site.sectors.order_by(Sector.code_sector.asc()).all()
    cells = []
    tech_set = set()
    antenna_models = set()
    for sector in sectors:
        sector_cells = sector.cells.order_by(Cell.cellname.asc()).all()
        cells.extend(sector_cells)
        for cell in sector_cells:
            tech_val = (cell.technology or "").strip().upper()
            if tech_val:
                tech_set.add(tech_val)
            if cell.antenna and cell.antenna.model:
                antenna_models.add(cell.antenna.model)

    # 3) Compute nearest sites from the same accessible perimeter.
    base_query = Site.query
    if accessible is not None:
        if not accessible:
            base_query = base_query.filter(False)
        else:
            base_query = base_query.filter(Site.id.in_(list(accessible)))
    candidates = base_query.filter(Site.id != site.id).all()

    nearest = []
    if site.latitude is not None and site.longitude is not None:
        for other in candidates:
            if other.latitude is None or other.longitude is None:
                continue
            dist_km = _haversine_km(site.latitude, site.longitude, other.latitude, other.longitude)
            nearest.append(
                {
                    "id": other.id,
                    "code_site": other.code_site,
                    "name": other.name,
                    "latitude": other.latitude,
                    "longitude": other.longitude,
                    "distance_km": round(dist_km, 2),
                }
            )
        nearest.sort(key=lambda row: row["distance_km"])
        nearest = nearest[:5]

    # 4) Return a single JSON payload consumed by the Site Profile modal.
    response = {
        "success": True,
        "site": {
            "id": site.id,
            "code_site": site.code_site,
            "name": site.name,
            "address": site.address,
            "latitude": site.latitude,
            "longitude": site.longitude,
            "altitude": site.altitude,
            "status": site.status,
            "support_nature": site.support_nature,
            "support_type": site.support_type,
            "support_height": site.support_height,
            "supplier_name": site.supplier.name if site.supplier else None,
            "commune_name": site.commune.name if site.commune else None,
            "wilaya_name": site.commune.wilaya.name if site.commune and site.commune.wilaya else None,
            "region_name": site.commune.wilaya.region.name if site.commune and site.commune.wilaya and site.commune.wilaya.region else None,
            "sectors_count": len(sectors),
            "cells_count": len(cells),
            "techs": sorted(tech_set),
            "antennas": sorted(antenna_models),
        },
        "sectors": [
            {
                "id": s.id,
                "code_sector": s.code_sector,
                "azimuth": s.azimuth,
                "hba": s.hba,
                "coverage_goal": s.coverage_goal,
                "cells_count": s.cells.count(),
            }
            for s in sectors
        ],
        "cells": [
            {
                "id": c.id,
                "cellname": c.cellname,
                "technology": c.technology,
                "frequency": c.frequency,
                "sector_code": c.sector.code_sector if c.sector else None,
                "antenna_model": c.antenna.model if c.antenna else None,
            }
            for c in cells
        ],
        "nearest_sites": nearest,
    }
    return jsonify(response)


@main_bp.route('/import-template/<entity>', methods=['GET'])
@login_required
def download_import_template(entity):
    key = (entity or "").strip().lower()
    if key in {"vendor", "vendors", "supplier"}:
        key = "suppliers"

    spec = IMPORT_TEMPLATE_SPECS.get(key)
    if not spec:
        flash(f'No template available for entity "{entity}".', "warning")
        return redirect(url_for("main.import_export"))

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in spec.get("sheets", {}).items():
        ws = wb.create_sheet(sheet_name[:31] or "Sheet1")
        ws.append(headers)
        sample_row = IMPORT_TEMPLATE_EXAMPLES.get(key, {}).get(sheet_name)
        if sample_row and len(sample_row) == len(headers):
            ws.append(sample_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{key}_import_template.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main_bp.route('/allplan_export')
@login_required
def allplan_export_page():
    return render_template('allplan_export.html', title='Allplan Export')


@main_bp.route('/lbs_export')
@login_required
def lbs_export_page():
    return render_template('lbs_export.html', title='LBS Export')


@main_bp.route('/kml_export')
@login_required
def kml_export_page():
    return render_template('kml_export.html', title='KML Export')


@main_bp.route('/export-data/<entity>', methods=['GET'])
@login_required
@admin_required
def export_data(entity):
    append_audit_event("export_data", entity, "SUCCESS", f"Export trigger for {entity}")
    flash(f"Démarrage de l'exportation pour l'entité \"{entity.upper()}\"...", "info")
    return redirect(url_for('main.import_export'))


@main_bp.route('/export-allplan', methods=['POST'])
@login_required
@csrf_protect
def export_allplan():
    uploaded_file = request.files.get("cell_file")
    raw_cells = request.form.get('cell_list', '')

    cell_names = []
    if uploaded_file and uploaded_file.filename:
        try:
            cell_names = _parse_cell_file(uploaded_file)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("main.import_export"))
    else:
        cell_names = _parse_cell_list(raw_cells)

    if not cell_names:
        flash('Veuillez fournir un fichier (ou une liste texte) contenant au moins une cellule.', 'warning')
        return redirect(url_for('main.import_export'))

    cells = (
        Cell.query.options(
            joinedload(Cell.sector).joinedload(Sector.site).joinedload(Site.commune),
            joinedload(Cell.antenna),
        )
        .filter(Cell.cellname.in_(cell_names))
        .all()
    )

    by_name = {c.cellname: c for c in cells}
    rows = {'2G': [], '3G': [], '4G': []}
    not_found_rows = []

    cell_codes = {_extract_cell_code(name) for name in cell_names}
    cell_codes.discard(None)
    mappings = Mapping.query.filter(Mapping.cell_code.in_(list(cell_codes))).all() if cell_codes else []

    mapping_full = {}
    mapping_tech = {}
    mapping_code = {}
    for m in mappings:
        code = (m.cell_code or "").strip()
        tech_key = (m.technology or "").strip().upper()
        band_key = (m.band or "").strip()
        if code and tech_key and band_key:
            mapping_full[(code, tech_key, band_key)] = m.sector_code
        if code and tech_key:
            mapping_tech[(code, tech_key)] = m.sector_code
        if code:
            mapping_code[code] = m.sector_code

    for name in cell_names:
        cell = by_name.get(name)
        if not cell:
            not_found_rows.append({
                "CELLNAME": name,
                "TECHNOLOGY": None,
                "ISSUE": "CELL_MISSING",
                "DETAIL": "Cellule introuvable dans la table Cell.",
            })
            continue
        tech = _normalize_tech(cell.technology)
        if tech not in rows:
            not_found_rows.append({
                "CELLNAME": cell.cellname,
                "TECHNOLOGY": cell.technology,
                "ISSUE": "TECHNOLOGY_UNSUPPORTED",
                "DETAIL": "Technologie non supportée pour export Allplan.",
            })
            continue

        sector = cell.sector
        site = sector.site if sector else None
        antenna = cell.antenna
        commune_name = site.commune.name if site and site.commune else None
        site_type = site.support_type if site else None
        cell_code = _extract_cell_code(cell.cellname)
        tech_key = (cell.technology or "").strip().upper()
        band_key = (cell.frequency or "").strip() if cell.frequency is not None else ""
        sector_id_from_mapping = (
            mapping_full.get((cell_code, tech_key, band_key))
            or mapping_tech.get((cell_code, tech_key))
            or mapping_code.get(cell_code)
        )

        issues = []
        if not sector:
            issues.append("SECTOR_MISSING")
        if sector and not site:
            issues.append("SITE_MISSING")
        if not sector_id_from_mapping:
            issues.append("MAPPING_MISSING")
        if issues:
            not_found_rows.append({
                "CELLNAME": cell.cellname,
                "TECHNOLOGY": cell.technology,
                "ISSUE": ", ".join(issues),
                "DETAIL": "Verifier les relations Cell->Sector->Site et la table Mapping.",
            })

        common = {
            'CELLNAME': cell.cellname,
            'LONGITUDE': site.longitude if site else None,
            'LATITUDE': site.latitude if site else None,
            'AZIMUTH': sector.azimuth if sector else None,
            'HEIGHT': sector.hba if sector else None,
            'ANTENNA': antenna.model if antenna else None,
            'ELECTRICALTILT': cell.tilt_electrical,
            'MECHANICALTILT': cell.tilt_mechanical,
            'HBEAMWIDTH': antenna.hbeamwidth if antenna else None,
            'VBEAMWIDTH': antenna.vbeamwidth if antenna else None,
            'GAIN': antenna.gain if antenna else None,
            'ANTENNAGAIN': antenna.gain if antenna else None,
            'SITENAME': site.code_site if site else None,
            'SITE_TYPE': site_type,
            'INDOORFLAG': None,
            'SECTORID': sector_id_from_mapping,
            'ARFCN': None,
            'DL UARFCN': None,
            'DL EARFCN': cell.frequency,
            'LAC': None,
            'CELLIDENTITY': None,
            'BCCH': None,
            'BSIC(octal)': None,
            'RNCID': None,
            'UTRANCELLIDENTITY': None,
            'ENODEBNAME': None,
            'ENODEBID': None,
            'CI': None,
            'PCI': None,
            'Core or Buffer': 'Core',
            'New Cell': None,
            'COMMUNE': commune_name,
        }
        rows[tech].append(common)

    wb = Workbook()
    wb.remove(wb.active)

    for tech in ('2G', '3G', '4G'):
        ws = wb.create_sheet(tech)
        headers = ALLPLAN_HEADERS[tech]
        ws.append(headers)
        for row in rows[tech]:
            ws.append([row.get(h) for h in headers])

    ws_nf = wb.create_sheet("NOT_FOUND")
    nf_headers = ["CELLNAME", "TECHNOLOGY", "ISSUE", "DETAIL"]
    ws_nf.append(nf_headers)
    for row in not_found_rows:
        ws_nf.append([row.get(h) for h in nf_headers])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"allplan_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@main_bp.route('/sync-cell-sectors', methods=['POST'])
@login_required
@csrf_protect
def sync_cell_sectors():
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    uploaded_file = request.files.get("cell_file")
    raw_cells = request.form.get("cell_list", "")
    scope = (request.form.get("scope") or "").strip().lower()
    search = (request.form.get("search") or "").strip()
    prioritized_cells = [c.strip() for c in request.form.getlist("prioritized_cells") if str(c).strip()]

    cell_names = []
    if uploaded_file and uploaded_file.filename:
        try:
            cell_names = _parse_cell_file(uploaded_file)
        except ValueError as exc:
            if is_ajax:
                return jsonify({"success": False, "message": str(exc)}), 400
            flash(str(exc), "danger")
            return redirect(request.referrer or url_for("main.import_export"))
    elif raw_cells.strip():
        cell_names = _parse_cell_list(raw_cells)

    if scope != "all" and not cell_names and not search and not prioritized_cells:
        msg = "Provide a cell file/list or select Sync ALL cells."
        if is_ajax:
            return jsonify({"success": False, "message": msg}), 400
        flash(msg, "warning")
        return redirect(request.referrer or url_for("main.import_export"))

    job_id = uuid.uuid4().hex
    append_audit_event("sync_start", "cells", "SUCCESS", f"Cell/Sector sync scope={scope}")
    _set_cell_sector_sync_job(
        job_id,
        status="queued",
        progress=0,
        message="Cell/Sector sync queued...",
        total=0,
        processed=0,
        updated=0,
        unchanged=0,
        unresolved=0,
        skipped=0,
        created_at=datetime.utcnow().isoformat(),
    )
    app_obj = current_app._get_current_object()
    t = threading.Thread(
        target=_run_cell_sector_sync_job,
        args=(app_obj, job_id, scope, cell_names, search, prioritized_cells),
        daemon=True,
    )
    t.start()

    status_url = url_for("main.cell_sector_sync_status", job_id=job_id)
    if is_ajax:
        return jsonify({"success": True, "job_id": job_id, "status_url": status_url}), 202

    flash("Cell/Sector sync started in background.", "info")
    return redirect(request.referrer or url_for("main.import_export"))


def _set_cell_sector_sync_job(job_id, **fields):
    with _cell_sector_sync_lock:
        job = _cell_sector_sync_jobs.get(job_id, {})
        job.update(fields)
        _cell_sector_sync_jobs[job_id] = job
        return dict(job)


def _get_cell_sector_sync_job(job_id):
    with _cell_sector_sync_lock:
        return dict(_cell_sector_sync_jobs.get(job_id, {}))


def _set_site_altitude_sync_job(job_id, **fields):
    with _site_altitude_sync_lock:
        job = _site_altitude_sync_jobs.get(job_id, {})
        job.update(fields)
        _site_altitude_sync_jobs[job_id] = job
        return dict(job)


def _get_site_altitude_sync_job(job_id):
    with _site_altitude_sync_lock:
        return dict(_site_altitude_sync_jobs.get(job_id, {}))


def _run_cell_sector_sync_job(app_obj, job_id, scope, cell_names, search, prioritized_cells):
    started_at = datetime.utcnow()
    _set_cell_sector_sync_job(
        job_id,
        status="processing",
        progress=1,
        message="Loading cells...",
        started_at=started_at.isoformat(),
    )

    try:
        from app.routes.import_data import resolve_sector_id_for_cell
        with app_obj.app_context():
            # Sync targets only cells without sector.
            query = Cell.query.filter(Cell.sector_id.is_(None))
            if scope != "all":
                if cell_names:
                    query = query.filter(Cell.cellname.in_(cell_names))
                elif search:
                    like = f"%{search}%"
                    query = query.filter(
                        or_(
                            Cell.cellname.ilike(like),
                            Cell.technology.ilike(like),
                            Cell.frequency.ilike(like),
                        )
                    )

            total = query.count()
            if total == 0:
                _set_cell_sector_sync_job(
                    job_id,
                    status="completed",
                    progress=100,
                    message="No cells found for sector sync.",
                    total=0,
                    processed=0,
                    finished_at=datetime.utcnow().isoformat(),
                )
                return

            _set_cell_sector_sync_job(job_id, total=total, message=f"Syncing {total} cells...")

            updated = 0
            unchanged = 0
            unresolved = 0
            skipped = 0
            processed = 0
            pending = 0
            batch_size = 2000

            priority_set = set(prioritized_cells or [])
            if priority_set:
                # Process currently visible rows first, then the rest.
                ordered = (
                    list(query.filter(Cell.cellname.in_(list(priority_set))).yield_per(1000))
                    + list(query.filter(~Cell.cellname.in_(list(priority_set))).yield_per(1000))
                )
            else:
                ordered = query.yield_per(1000)

            for cell in ordered:
                tech = (cell.technology or "").strip().upper()
                freq = (cell.frequency or "").strip()
                if not cell.cellname or not tech or not freq:
                    skipped += 1
                else:
                    sector_id, _ = resolve_sector_id_for_cell(cell.cellname, tech, freq)
                    if sector_id is None:
                        unresolved += 1
                    elif cell.sector_id == int(sector_id):
                        unchanged += 1
                    else:
                        cell.sector_id = int(sector_id)
                        updated += 1
                        pending += 1

                processed += 1
                if pending >= batch_size:
                    db.session.commit()
                    pending = 0

                if processed == 1 or processed % 500 == 0 or processed == total:
                    pct = int(processed * 100 / total)
                    _set_cell_sector_sync_job(
                        job_id,
                        progress=pct,
                        processed=processed,
                        updated=updated,
                        unchanged=unchanged,
                        unresolved=unresolved,
                        skipped=skipped,
                        message=f"Sync {processed}/{total}",
                    )

            db.session.commit()
            finished_at = datetime.utcnow()
            _set_cell_sector_sync_job(
                job_id,
                status="completed",
                progress=100,
                processed=processed,
                updated=updated,
                unchanged=unchanged,
                unresolved=unresolved,
                skipped=skipped,
                finished_at=finished_at.isoformat(),
                message=(
                    f"Cell sector sync done: {updated} updated, {unchanged} unchanged, "
                    f"{unresolved} unresolved, {skipped} skipped."
                ),
            )
    except Exception as exc:
        with app_obj.app_context():
            db.session.rollback()
        _set_cell_sector_sync_job(
            job_id,
            status="failed",
            progress=100,
            message=f"Cell/Sector sync failed: {exc}",
            finished_at=datetime.utcnow().isoformat(),
        )


@main_bp.route('/sync-cell-sectors/status/<job_id>', methods=['GET'])
@login_required
def cell_sector_sync_status(job_id):
    job = _get_cell_sector_sync_job(job_id)
    if not job:
        return jsonify({"success": False, "message": "Sync job not found."}), 404

    return jsonify({
        "success": True,
        "job_id": job_id,
        "status": job.get("status", "unknown"),
        "progress": int(job.get("progress", 0)),
        "message": job.get("message", ""),
        "total": int(job.get("total", 0)),
        "processed": int(job.get("processed", 0)),
        "updated": int(job.get("updated", 0)),
        "unchanged": int(job.get("unchanged", 0)),
        "unresolved": int(job.get("unresolved", 0)),
        "skipped": int(job.get("skipped", 0)),
    }), 200


@main_bp.route('/sync-site-altitudes', methods=['POST'])
@login_required
@csrf_protect
def sync_site_altitudes():
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    scope = (request.form.get("scope") or "").strip().lower()
    search = (request.form.get("search") or "").strip()

    prioritized_sites = []
    for raw in request.form.getlist("prioritized_sites"):
        try:
            prioritized_sites.append(int(raw))
        except (TypeError, ValueError):
            continue
    prioritized_sites = sorted(set(prioritized_sites))

    if scope != "all" and not search and not prioritized_sites:
        msg = "Provide a filter or select Sync ALL sites."
        if is_ajax:
            return jsonify({"success": False, "message": msg}), 400
        flash(msg, "warning")
        return redirect(request.referrer or url_for("list_bp.view_sites"))

    job_id = uuid.uuid4().hex
    append_audit_event("sync_start", "sites", "SUCCESS", f"Site altitude sync scope={scope}")
    accessible_sites_snapshot = get_accessible_site_ids()
    _set_site_altitude_sync_job(
        job_id,
        status="queued",
        progress=0,
        message="Site altitude sync queued...",
        total=0,
        processed=0,
        updated=0,
        unresolved=0,
        skipped=0,
        created_at=datetime.utcnow().isoformat(),
    )

    app_obj = current_app._get_current_object()
    t = threading.Thread(
        target=_run_site_altitude_sync_job,
        args=(app_obj, job_id, scope, search, prioritized_sites, accessible_sites_snapshot),
        daemon=True,
    )
    t.start()

    status_url = url_for("main.site_altitude_sync_status", job_id=job_id)
    if is_ajax:
        return jsonify({"success": True, "job_id": job_id, "status_url": status_url}), 202

    flash("Site altitude sync started in background.", "info")
    return redirect(request.referrer or url_for("list_bp.view_sites"))


def _run_site_altitude_sync_job(app_obj, job_id, scope, search, prioritized_sites, accessible_sites_snapshot):
    _set_site_altitude_sync_job(
        job_id,
        status="processing",
        progress=1,
        message="Loading sites...",
        started_at=datetime.utcnow().isoformat(),
    )
    try:
        with app_obj.app_context():
            query = Site.query
            accessible_sites = accessible_sites_snapshot
            if accessible_sites is not None:
                if not accessible_sites:
                    query = query.filter(False)
                else:
                    query = query.filter(Site.id.in_(list(accessible_sites)))

            # Target only sites missing altitude.
            query = query.filter(Site.altitude.is_(None))

            if scope != "all":
                if prioritized_sites:
                    query = query.filter(Site.id.in_(prioritized_sites))
                elif search:
                    like = f"%{search}%"
                    query = query.filter(or_(Site.code_site.ilike(like), Site.name.ilike(like)))

            total = query.count()
            if total == 0:
                _set_site_altitude_sync_job(
                    job_id,
                    status="completed",
                    progress=100,
                    message="No sites to update (altitude already filled or no match).",
                    total=0,
                    processed=0,
                    finished_at=datetime.utcnow().isoformat(),
                )
                return

            _set_site_altitude_sync_job(job_id, total=total, message=f"Syncing altitude for {total} sites...")

            updated = 0
            unresolved = 0
            skipped = 0
            processed = 0
            pending = 0
            batch_size = 100

            ordered = query.yield_per(200)
            for site in ordered:
                lat = site.latitude
                lon = site.longitude
                if lat is None or lon is None:
                    skipped += 1
                else:
                    alt = _fetch_ground_altitude(lat, lon)
                    if alt is None:
                        unresolved += 1
                    else:
                        site.altitude = round(float(alt), 1)
                        updated += 1
                        pending += 1

                processed += 1
                if pending >= batch_size:
                    db.session.commit()
                    pending = 0

                if processed == 1 or processed % 25 == 0 or processed == total:
                    pct = int(processed * 100 / total)
                    _set_site_altitude_sync_job(
                        job_id,
                        progress=pct,
                        processed=processed,
                        updated=updated,
                        unresolved=unresolved,
                        skipped=skipped,
                        message=f"Altitude sync {processed}/{total}",
                    )

            db.session.commit()
            _set_site_altitude_sync_job(
                job_id,
                status="completed",
                progress=100,
                processed=processed,
                updated=updated,
                unresolved=unresolved,
                skipped=skipped,
                finished_at=datetime.utcnow().isoformat(),
                message=(
                    f"Site altitude sync done: {updated} updated, "
                    f"{unresolved} unresolved, {skipped} skipped."
                ),
            )
    except Exception as exc:
        with app_obj.app_context():
            db.session.rollback()
        _set_site_altitude_sync_job(
            job_id,
            status="failed",
            progress=100,
            message=f"Site altitude sync failed: {exc}",
            finished_at=datetime.utcnow().isoformat(),
        )


@main_bp.route('/sync-site-altitudes/status/<job_id>', methods=['GET'])
@login_required
def site_altitude_sync_status(job_id):
    job = _get_site_altitude_sync_job(job_id)
    if not job:
        return jsonify({"success": False, "message": "Sync job not found."}), 404

    return jsonify({
        "success": True,
        "job_id": job_id,
        "status": job.get("status", "unknown"),
        "progress": int(job.get("progress", 0)),
        "message": job.get("message", ""),
        "total": int(job.get("total", 0)),
        "processed": int(job.get("processed", 0)),
        "updated": int(job.get("updated", 0)),
        "unresolved": int(job.get("unresolved", 0)),
        "skipped": int(job.get("skipped", 0)),
    }), 200


@main_bp.route('/export-lbs', methods=['POST'])
@login_required
@csrf_protect
def export_lbs():
    uploaded_file = request.files.get("cell_file")
    raw_cells = request.form.get('cell_list', '')

    cell_names = []
    if uploaded_file and uploaded_file.filename:
        try:
            cell_names = _parse_cell_file(uploaded_file)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("main.lbs_export_page"))
    else:
        cell_names = _parse_cell_list(raw_cells)

    if not cell_names:
        flash("Please provide a non-empty cell list (.xlsx/.csv or text).", "warning")
        return redirect(url_for("main.lbs_export_page"))

    ran_map = build_ran_reference_map(current_app.instance_path)

    cells = (
        Cell.query.options(
            joinedload(Cell.sector).joinedload(Sector.site).joinedload(Site.commune),
            joinedload(Cell.antenna),
            joinedload(Cell.profile_2g),
            joinedload(Cell.profile_3g),
            joinedload(Cell.profile_4g),
        )
        .filter(Cell.cellname.in_(cell_names))
        .all()
    )

    by_name = {c.cellname: c for c in cells}
    rows = {"2G": [], "3G": [], "4G": []}
    not_found_rows = []

    for name in cell_names:
        cell = by_name.get(name)
        if not cell:
            not_found_rows.append({
                "CELLNAME": name,
                "TECHNOLOGY": None,
                "ISSUE": "CELL_MISSING",
                "DETAIL": "Cell not found in Cell table.",
            })
            continue

        tech = _normalize_tech(cell.technology)
        if tech not in rows:
            not_found_rows.append({
                "CELLNAME": cell.cellname,
                "TECHNOLOGY": cell.technology,
                "ISSUE": "TECHNOLOGY_UNSUPPORTED",
                "DETAIL": "Supported technologies for LBS export are 2G/3G/4G.",
            })
            continue

        sector = cell.sector
        site = sector.site if sector else None
        antenna = cell.antenna
        cid = _cell_suffix_int(cell.cellname)
        band = _freq_band_int(cell.frequency)
        common = {
            "site_code": site.code_site if site else None,
            "cellname": cell.cellname,
            "lon": site.longitude if site else None,
            "lat": site.latitude if site else None,
            "ant_type": _antenna_type_for_lbs(site),
            "radius": _lbs_radius_km(tech, band, ran_map),
            "gain": antenna.gain if antenna else None,
            "pwr": _lbs_nominal_power_dbm(tech, band, ran_map),
            "spec": antenna.model if antenna else None,
            "hba": sector.hba if sector else None,
            "down_tilt": cell.tilt_electrical,
            "mech_tilt": cell.tilt_mechanical,
            "elec_tilt": cell.tilt_electrical,
            "az": sector.azimuth if sector else None,
            "h_bw": antenna.hbeamwidth if antenna else None,
            "band": band,
        }

        issues = []
        if not sector:
            issues.append("SECTOR_MISSING")
        if sector and not site:
            issues.append("SITE_MISSING")

        if tech == "2G":
            p2 = cell.profile_2g
            ncc, bcc = _ncc_bcc_from_bsic(p2.bsic if p2 else None)
            rows["2G"].append({
                "CID": p2.ci if p2 and p2.ci is not None else cid,
                "Site Name": common["site_code"],
                "CellName": common["cellname"],
                "Longitude": common["lon"],
                "Latitude": common["lat"],
                "AntennaType": common["ant_type"],
                "MaxCellRadius": common["radius"],
                "AntennaGain": common["gain"],
                "BSNominalPower": common["pwr"],
                "BSPowerBCCH": 20,
                "TAlimit": 5,
                "AntennaSpec": common["spec"],
                "HeightAGL": common["hba"],
                "DownTilt": common["down_tilt"],
                "MechanicalTilt": common["mech_tilt"],
                "ElectricalTilt": common["elec_tilt"],
                "Azimuth": common["az"],
                "HorizBeamWidth": common["h_bw"],
                "LAC": p2.lac if p2 else None,
                "NCC": ncc,
                "BCC": bcc,
                "BCCH": p2.bcch if p2 else None,
                "MCC": 603,
                "MNC": 2,
                "FrequencyBand": common["band"],
                "Node Name": p2.bsc if p2 else None,
            })
        elif tech == "3G":
            p3 = cell.profile_3g
            rows["3G"].append({
                "CID": p3.ci if p3 and p3.ci is not None else cid,
                "SiteName": common["site_code"],
                "CellName": common["cellname"],
                "Longitude": common["lon"],
                "Latitude": common["lat"],
                "AntennaType": common["ant_type"],
                "MaxCellRadius": common["radius"],
                "Gain": common["gain"],
                "BSNominalPower": common["pwr"],
                "AntennaSpec": common["spec"],
                "HeightAGL": common["hba"],
                "DownTilt": common["down_tilt"],
                "Azimuth": common["az"],
                "MechanicalTilt": common["mech_tilt"],
                "ElectricalTilt": common["elec_tilt"],
                "HorizBeamWidth": common["h_bw"],
                "LAC": p3.lac if p3 else None,
                "SAC": cid,
                "RNCid": p3.rnc if p3 else None,
                "PSC": p3.psc if p3 else None,
                "UARFCN": p3.dlarfcn if p3 else None,
                "FrequencyBand": common["band"],
                "NodeName": p3.rnc if p3 else None,
            })
        else:
            p4 = cell.profile_4g
            rows["4G"].append({
                "Site Name": common["site_code"],
                "CellName": common["cellname"],
                "Longitude": common["lon"],
                "Latitude": common["lat"],
                "AntennaType": common["ant_type"],
                "MaxCellRadius": common["radius"],
                "AntennaGain": common["gain"],
                "BSNominalPower": common["pwr"],
                "AntennaSpec": common["spec"],
                "HeightAGL": common["hba"],
                "DownTilt": common["down_tilt"],
                "MechanicalTilt": common["mech_tilt"],
                "ElectricalTilt": common["elec_tilt"],
                "Azimuth": common["az"],
                "HorizBeamWidth": common["h_bw"],
                "TAC ": p4.tac if p4 else None,
                "E-UTRANCELLID": p4.ci if p4 and p4.ci is not None else cid,
                "EARFCN": p4.earfcn if p4 else None,
                "FrequencyBand": common["band"],
                "PCI": p4.pci if p4 else None,
                "NodeName": p4.enodeb if p4 else None,
            })

        if issues:
            not_found_rows.append({
                "CELLNAME": cell.cellname,
                "TECHNOLOGY": cell.technology,
                "ISSUE": ", ".join(issues),
                "DETAIL": "Missing relationship Cell->Sector->Site.",
            })

    wb = Workbook()
    wb.remove(wb.active)
    for tech in ("2G", "3G", "4G"):
        ws = wb.create_sheet(tech)
        headers = LBS_HEADERS[tech]
        ws.append(headers)
        for row in rows[tech]:
            ws.append([row.get(h) for h in headers])

    ws_nf = wb.create_sheet("NOT_FOUND")
    nf_headers = ["CELLNAME", "TECHNOLOGY", "ISSUE", "DETAIL"]
    ws_nf.append(nf_headers)
    for row in not_found_rows:
        ws_nf.append([row.get(h) for h in nf_headers])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"lbs_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
