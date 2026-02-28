import io
import logging
import os
import threading
import uuid
from datetime import datetime
from math import asin, atan2, cos, radians, sin, degrees
from pathlib import Path
from xml.sax.saxutils import escape

from flask import Blueprint, abort, current_app, jsonify, request, send_file, url_for
from openpyxl import load_workbook
from sqlalchemy.orm import joinedload

from app.models import Cell, Commune, Sector, Site, Wilaya
from app.security import csrf_protect, get_accessible_site_ids, login_required

doc_bp = Blueprint('doc_bp', __name__)
logger = logging.getLogger(__name__)
_kml_jobs = {}
_kml_jobs_lock = threading.Lock()


def _site_allowed(site):
    accessible_sites = get_accessible_site_ids()
    if accessible_sites is None:
        return True
    return site.id in accessible_sites


def _set_kml_job(job_id, **fields):
    with _kml_jobs_lock:
        job = _kml_jobs.get(job_id, {})
        job.update(fields)
        _kml_jobs[job_id] = job
        return dict(job)


def _get_kml_job(job_id):
    with _kml_jobs_lock:
        return dict(_kml_jobs.get(job_id, {}))


def _kml_exports_dir():
    out_dir = Path(current_app.instance_path) / "kml_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _kml_document(placemarks):
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<kml xmlns="http://www.opengis.net/kml/2.2">\n'
        '  <Document>\n'
        f"{placemarks}"
        '  </Document>\n'
        '</kml>\n'
    )


def _site_icon_href(icon_key):
    key = (icon_key or "tower").strip().lower()
    icon_map = {
        "tower": "http://maps.google.com/mapfiles/kml/shapes/placemark_square.png",
        "antenna": "http://maps.google.com/mapfiles/kml/shapes/radio.png",
        "target": "http://maps.google.com/mapfiles/kml/shapes/target.png",
        "pin_blue": "http://maps.google.com/mapfiles/kml/paddle/blu-circle.png",
        "pin_red": "http://maps.google.com/mapfiles/kml/paddle/red-circle.png",
        "pin_green": "http://maps.google.com/mapfiles/kml/paddle/grn-circle.png",
        "pin_yellow": "http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png",
        "pin_white": "http://maps.google.com/mapfiles/kml/paddle/wht-circle.png",
        "circle": "http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png",
    }
    return icon_map.get(key, icon_map["tower"])


def _build_site_placemark(site, icon_href, icon_scale):
    lon = site.longitude
    lat = site.latitude
    if lon is None or lat is None:
        return ''

    description_html = (
        "<![CDATA["
        "<div style='margin:0;padding:0;font-family:Segoe UI,Arial,sans-serif;font-size:12px;color:#0f172a;min-width:320px;'>"
        f"<div style='background:#0f172a;color:#f8fafc;font-size:18px;font-weight:700;letter-spacing:0.3px;padding:10px 12px;border-radius:10px 10px 0 0;'>{escape(site.code_site or '')}</div>"
        "<table style='width:100%;border-collapse:separate;border-spacing:0;border:1px solid #d1d5db;border-top:0;border-radius:0 0 10px 10px;overflow:hidden;'>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;width:42%;border-top:1px solid #e5e7eb;'>SITE NAME</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(site.name or '')}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>ALTITUDE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.altitude or ''))}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>VENDOR</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(site.supplier.name if site.supplier else '')}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>SUPPORT NATURE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.support_nature or ''))}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>SUPPORT TYPE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.support_type or ''))}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>SUPPORT HEIGHT</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.support_height or ''))}</td></tr>"
        "</table>"
        "</div>"
        "]]>"
    )

    return (
        '    <Placemark>\n'
        f'      <name>{escape(site.code_site or "Site")}</name>\n'
        '      <Snippet maxLines="0"></Snippet>\n'
        f'      <description>{description_html}</description>\n'
        '      <Style>\n'
        f'        <IconStyle><scale>{icon_scale:.1f}</scale><Icon><href>{icon_href}</href></Icon></IconStyle>\n'
        '        <BalloonStyle><text>$[description]</text></BalloonStyle>\n'
        '      </Style>\n'
        '      <Point>\n'
        f'        <coordinates>{lon},{lat},0</coordinates>\n'
        '      </Point>\n'
        '    </Placemark>\n'
    )


def _build_sector_placemark(sector):
    site = sector.site
    if not site or site.longitude is None or site.latitude is None:
        return ''

    cells = sector.cells.all() if hasattr(sector.cells, "all") else list(sector.cells)
    antenna_models = sorted({cell.antenna.model for cell in cells if getattr(cell, 'antenna', None) and cell.antenna.model})
    antenna_text = ', '.join(antenna_models) if antenna_models else 'N/A'
    beamwidth = _sector_beamwidth(sector, cells)
    beam_polygon = _sector_beam_polygon(sector, site, beamwidth=beamwidth, radius_km=0.8, points=20)

    description_html = (
        "<![CDATA["
        "<div style='font-family:Arial,sans-serif;font-size:12px;'>"
        "<h4 style='margin:0 0 6px 0;color:#0f4c81;'>Sector Beam</h4>"
        "<table style='border-collapse:collapse;'>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>Site</b></td><td>{escape(site.code_site or '')}</td></tr>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>Sector</b></td><td>{escape(sector.code_sector or '')}</td></tr>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>Azimuth</b></td><td>{escape(str(sector.azimuth or ''))} deg</td></tr>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>Beamwidth</b></td><td>{escape(str(beamwidth))} deg</td></tr>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>HBA</b></td><td>{escape(str(sector.hba or ''))}</td></tr>"
        f"<tr><td style='padding:2px 8px 2px 0;'><b>Antennas</b></td><td>{escape(antenna_text)}</td></tr>"
        "</table>"
        "</div>"
        "]]>"
    )

    style = (
        "      <Style>\n"
        "        <LineStyle><color>ff0055ff</color><width>1.2</width></LineStyle>\n"
        "        <PolyStyle><color>550055ff</color></PolyStyle>\n"
        "      </Style>\n"
    )

    return (
        '    <Placemark>\n'
        f'      <name>{escape(sector.code_sector or "Sector")}</name>\n'
        f'      <description>{description_html}</description>\n'
        f"{style}"
        '      <Polygon>\n'
        '        <outerBoundaryIs><LinearRing><coordinates>\n'
        f'          {beam_polygon}\n'
        '        </coordinates></LinearRing></outerBoundaryIs>\n'
        '      </Polygon>\n'
        '    </Placemark>\n'
    )


def _sector_beamwidth(sector, cells):
    widths = []
    for cell in cells:
        antenna = getattr(cell, "antenna", None)
        hbw = getattr(antenna, "hbeamwidth", None)
        if hbw is None:
            continue
        try:
            hbw = float(hbw)
        except (TypeError, ValueError):
            continue
        if 2 <= hbw <= 180:
            widths.append(hbw)

    if widths:
        return round(sum(widths) / len(widths), 1)
    return 40.0


def _destination(lat, lon, bearing_deg, distance_km):
    r = 6371.0
    brng = radians(bearing_deg)
    lat1 = radians(lat)
    lon1 = radians(lon)
    d = distance_km / r

    lat2 = asin(sin(lat1) * cos(d) + cos(lat1) * sin(d) * cos(brng))
    lon2 = lon1 + atan2(
        sin(brng) * sin(d) * cos(lat1),
        cos(d) - sin(lat1) * sin(lat2),
    )
    return degrees(lat2), degrees(lon2)


def _sector_beam_polygon(sector, site, beamwidth=40.0, radius_km=0.8, points=20):
    try:
        azimuth = float(getattr(sector, "azimuth", 0.0) or 0.0)
    except (TypeError, ValueError, AttributeError):
        azimuth = 0.0

    half = beamwidth / 2.0
    coords = [(site.longitude, site.latitude, 0)]

    for i in range(points + 1):
        angle = azimuth - half + (beamwidth * i / points)
        lat2, lon2 = _destination(site.latitude, site.longitude, angle, radius_km)
        coords.append((lon2, lat2, 0))

    coords.append((site.longitude, site.latitude, 0))
    return " ".join(f"{lon},{lat},{alt}" for lon, lat, alt in coords)


def _clamp(value, lo, hi):
    return max(lo, min(hi, value))


def _parse_hex_color(color_value):
    raw = (color_value or "").strip().lstrip("#")
    if len(raw) == 3:
        raw = "".join(ch * 2 for ch in raw)
    if len(raw) != 6:
        return "0055ff"
    try:
        int(raw, 16)
    except ValueError:
        return "0055ff"
    return raw.lower()


def _kml_color_from_rgb(hex_rgb, alpha_hex):
    rr = hex_rgb[0:2]
    gg = hex_rgb[2:4]
    bb = hex_rgb[4:6]
    return f"{alpha_hex}{bb}{gg}{rr}"


@doc_bp.route('/generate_d4b/<int:site_id>')
@login_required
def generate_d4b(site_id):
    site = Site.query.get_or_404(site_id)
    if not _site_allowed(site):
        abort(403, description='Acces refuse a ce site.')

    sectors = (
        Sector.query.filter_by(site_id=site_id)
        .order_by(Sector.code_sector)
        .all()
    )

    template_path = os.path.join('app', 'static', 'template_D4b.xlsx')

    if not os.path.exists(template_path):
        abort(404, description='Template D4b introuvable dans app/static/')

    wb = load_workbook(template_path)
    ws = wb.active

    def get_sector_value(index, attr):
        if len(sectors) > index:
            val = getattr(sectors[index], attr)
            return val if val is not None else ''
        return ''

    def get_sector_antennas(index):
        if len(sectors) <= index:
            return ''
        sector = sectors[index]
        cells = Cell.query.filter_by(sector_id=sector.id).all()
        models = sorted({c.antenna.model for c in cells if c.antenna and c.antenna.model})
        return ' / '.join(models)

    try:
        ws['J5'] = site.code_site or ''
        ws['C5'] = site.name or ''
        ws['J9'] = f"{site.code_site} C1" if site.code_site else 'C1'

        ws['E11'].value = site.address or ''
        ws['H18'].value = site.longitude or ''
        ws['J18'].value = site.latitude or ''

        support_parts = [
            str(site.support_nature or ''),
            str(site.support_type or ''),
            f"{site.support_height} m" if site.support_height else '',
        ]
        ws['H23'].value = ' '.join(filter(None, support_parts))

        ws['E31'] = get_sector_value(0, 'azimuth')
        ws['G31'] = get_sector_value(1, 'azimuth')
        ws['I31'] = get_sector_value(2, 'azimuth')
        ws['K31'] = get_sector_value(3, 'azimuth')

        ws['E32'] = get_sector_value(0, 'hba')
        ws['G32'] = get_sector_value(1, 'hba')
        ws['I32'] = get_sector_value(2, 'hba')
        ws['K32'] = get_sector_value(3, 'hba')

        # Antenna values extracted from cells (a sector can host multiple antennas)
        ws['E33'] = get_sector_antennas(0)
        ws['G33'] = get_sector_antennas(1)
        ws['I33'] = get_sector_antennas(2)
        ws['K33'] = get_sector_antennas(3)

        ws['C42'] = get_sector_value(0, 'coverage_goal')
        ws['C43'] = get_sector_value(1, 'coverage_goal')
        ws['C44'] = get_sector_value(2, 'coverage_goal')
        ws['C45'] = get_sector_value(3, 'coverage_goal')

    except Exception:
        logger.exception('Erreur lors du remplissage du D4b')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"D4b_{site.code_site}.xlsx",
    )


@doc_bp.route('/export_kml/sites')
@login_required
def export_kml_sites():
    icon_href = _site_icon_href(request.args.get('site_icon', default='tower', type=str))
    icon_scale = _clamp(request.args.get('site_icon_scale', default=1.2, type=float) or 1.2, 0.8, 1.8)
    content = _build_sites_kml_content(
        icon_href=icon_href,
        icon_scale=icon_scale,
        region_id=_safe_int(request.args.get("region_id")),
        wilaya_id=_safe_int(request.args.get("wilaya_id")),
        commune_id=_safe_int(request.args.get("commune_id")),
        site_id=_safe_int(request.args.get("site_id")),
    )

    return send_file(
        io.BytesIO(content.encode('utf-8')),
        as_attachment=True,
        download_name='sites_export.kml',
        mimetype='application/vnd.google-earth.kml+xml',
    )


@doc_bp.route('/export_kml/sectors')
@login_required
def export_kml_sectors():
    beam_length_km = _clamp(request.args.get('beam_length_km', default=0.8, type=float) or 0.8, 0.1, 10.0)
    beam_width_deg = _clamp(request.args.get('beam_width_deg', default=40.0, type=float) or 40.0, 5.0, 180.0)
    beam_rgb = _parse_hex_color(request.args.get('beam_color', default='#0055ff', type=str))
    line_color = _kml_color_from_rgb(beam_rgb, "ff")
    poly_color = _kml_color_from_rgb(beam_rgb, "66")

    content = _build_sectors_kml_content(
        beam_length_km=beam_length_km,
        beam_width_deg=beam_width_deg,
        line_color=line_color,
        poly_color=poly_color,
        region_id=_safe_int(request.args.get("region_id")),
        wilaya_id=_safe_int(request.args.get("wilaya_id")),
        commune_id=_safe_int(request.args.get("commune_id")),
        site_id=_safe_int(request.args.get("site_id")),
    )

    return send_file(
        io.BytesIO(content.encode('utf-8')),
        as_attachment=True,
        download_name='sectors_export.kml',
        mimetype='application/vnd.google-earth.kml+xml',
    )


def _safe_int(value):
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _iter_accessible_sites(region_id=None, wilaya_id=None, commune_id=None, site_id=None):
    query = Site.query.order_by(Site.code_site.asc())
    accessible_sites = get_accessible_site_ids()
    if accessible_sites is not None:
        if not accessible_sites:
            return []
        query = query.filter(Site.id.in_(list(accessible_sites)))
    if site_id is not None:
        query = query.filter(Site.id == int(site_id))
    if commune_id is not None:
        query = query.filter(Site.commune_id == int(commune_id))
    if wilaya_id is not None:
        query = query.filter(Site.commune.has(wilaya_id=int(wilaya_id)))
    if region_id is not None:
        query = query.filter(Site.commune.has(Commune.wilaya.has(region_id=int(region_id))))
    return query


def _iter_accessible_sectors(region_id=None, wilaya_id=None, commune_id=None, site_id=None):
    query = (
        Sector.query
        .join(Site, Sector.site_id == Site.id)
        .options(joinedload(Sector.site).joinedload(Site.commune))
        .order_by(Sector.code_sector.asc())
    )
    accessible_sites = get_accessible_site_ids()
    if accessible_sites is not None:
        if not accessible_sites:
            return []
        query = query.filter(Site.id.in_(list(accessible_sites)))
    if site_id is not None:
        query = query.filter(Sector.site_id == int(site_id))
    if commune_id is not None:
        query = query.filter(Site.commune_id == int(commune_id))
    if wilaya_id is not None:
        query = query.filter(Site.commune.has(wilaya_id=int(wilaya_id)))
    if region_id is not None:
        query = query.filter(Site.commune.has(Commune.wilaya.has(region_id=int(region_id))))
    return query


def _build_sites_kml_content(icon_href, icon_scale, progress_cb=None, region_id=None, wilaya_id=None, commune_id=None, site_id=None):
    query = _iter_accessible_sites(region_id=region_id, wilaya_id=wilaya_id, commune_id=commune_id, site_id=site_id)
    if isinstance(query, list):
        abort(403, description='Aucun site autorise pour cet utilisateur.')

    total = query.count()
    placemarks = []
    processed = 0
    for site in query.yield_per(1000):
        placemarks.append(_build_site_placemark(site, icon_href=icon_href, icon_scale=icon_scale))
        processed += 1
        if progress_cb and (processed == 1 or processed % 300 == 0 or processed == total):
            progress_cb(processed, total, f"Building sites KML {processed}/{total}")
    return _kml_document(''.join(placemarks))


def _prefetch_cells_by_sector(sector_ids, progress_cb=None):
    cells_by_sector = {}
    total_batches = max((len(sector_ids) + 999) // 1000, 1)
    for idx in range(total_batches):
        chunk = sector_ids[idx * 1000:(idx + 1) * 1000]
        if not chunk:
            continue
        cells = (
            Cell.query
            .options(joinedload(Cell.antenna))
            .filter(Cell.sector_id.in_(chunk))
            .all()
        )
        for cell in cells:
            cells_by_sector.setdefault(cell.sector_id, []).append(cell)
        if progress_cb:
            progress_cb(idx + 1, total_batches, f"Loading sector cells {idx + 1}/{total_batches}")
    return cells_by_sector


def _build_sectors_kml_content(beam_length_km, beam_width_deg, line_color, poly_color, progress_cb=None, region_id=None, wilaya_id=None, commune_id=None, site_id=None):
    query = _iter_accessible_sectors(region_id=region_id, wilaya_id=wilaya_id, commune_id=commune_id, site_id=site_id)
    if isinstance(query, list):
        abort(403, description='Aucun secteur autorise pour cet utilisateur.')

    sectors = query.all()
    if not sectors:
        return _kml_document("")

    sector_ids = [s.id for s in sectors]
    cells_by_sector = _prefetch_cells_by_sector(
        sector_ids,
        progress_cb=(lambda done, total, msg: progress_cb(done, total * 2, msg) if progress_cb else None),
    )

    total = len(sectors)
    placemarks = []
    for idx, sector in enumerate(sectors, start=1):
        placemarks.append(
            _build_sector_placemark_with_options(
                sector,
                beam_length_km=beam_length_km,
                beam_width_deg=beam_width_deg,
                line_color=line_color,
                poly_color=poly_color,
                preloaded_cells=cells_by_sector.get(sector.id, []),
            )
        )
        if progress_cb and (idx == 1 or idx % 300 == 0 or idx == total):
            progress_cb(total + idx, total * 2, f"Building sectors KML {idx}/{total}")

    return _kml_document(''.join(placemarks))


def _run_kml_job(app_obj, job_id, kind, params):
    started_at = datetime.utcnow()
    _set_kml_job(job_id, status="processing", progress=2, message="Starting KML export...", started_at=started_at.isoformat())
    try:
        with app_obj.app_context():
            region_id = _safe_int(params.get("region_id"))
            wilaya_id = _safe_int(params.get("wilaya_id"))
            commune_id = _safe_int(params.get("commune_id"))
            site_id = _safe_int(params.get("site_id"))
            if kind == "sites":
                icon_href = _site_icon_href(params.get("site_icon"))
                icon_scale = _clamp(float(params.get("site_icon_scale", 1.2)), 0.8, 1.8)
                content = _build_sites_kml_content(
                    icon_href=icon_href,
                    icon_scale=icon_scale,
                    region_id=region_id,
                    wilaya_id=wilaya_id,
                    commune_id=commune_id,
                    site_id=site_id,
                    progress_cb=lambda done, total, msg: _set_kml_job(
                        job_id,
                        progress=int((done * 100 / max(total, 1))),
                        processed=int(done),
                        total=int(total),
                        message=msg,
                    ),
                )
                download_name = "sites_export.kml"
            else:
                beam_length_km = _clamp(float(params.get("beam_length_km", 0.8)), 0.1, 10.0)
                beam_width_deg = _clamp(float(params.get("beam_width_deg", 40.0)), 5.0, 180.0)
                beam_rgb = _parse_hex_color(params.get("beam_color", "#0055ff"))
                line_color = _kml_color_from_rgb(beam_rgb, "ff")
                poly_color = _kml_color_from_rgb(beam_rgb, "66")
                content = _build_sectors_kml_content(
                    beam_length_km=beam_length_km,
                    beam_width_deg=beam_width_deg,
                    line_color=line_color,
                    poly_color=poly_color,
                    region_id=region_id,
                    wilaya_id=wilaya_id,
                    commune_id=commune_id,
                    site_id=site_id,
                    progress_cb=lambda done, total, msg: _set_kml_job(
                        job_id,
                        progress=int((done * 100 / max(total, 1))),
                        processed=int(done),
                        total=int(total),
                        message=msg,
                    ),
                )
                download_name = "sectors_export.kml"

            out_path = _kml_exports_dir() / f"kml_{job_id}.kml"
            out_path.write_text(content, encoding="utf-8")
            _set_kml_job(
                job_id,
                status="completed",
                progress=100,
                message="KML export completed.",
                file_path=str(out_path),
                download_name=download_name,
                finished_at=datetime.utcnow().isoformat(),
            )
    except Exception as exc:
        logger.exception("KML async export failed")
        _set_kml_job(
            job_id,
            status="failed",
            progress=100,
            message=f"KML export failed: {exc}",
            finished_at=datetime.utcnow().isoformat(),
        )


@doc_bp.route('/export_kml/sites/start', methods=['POST'])
@login_required
@csrf_protect
def start_kml_sites_export():
    job_id = uuid.uuid4().hex
    params = {
        "site_icon": request.form.get("site_icon", "tower"),
        "site_icon_scale": request.form.get("site_icon_scale", "1.2"),
        "region_id": request.form.get("region_id", ""),
        "wilaya_id": request.form.get("wilaya_id", ""),
        "commune_id": request.form.get("commune_id", ""),
        "site_id": request.form.get("site_id", ""),
    }
    _set_kml_job(job_id, status="queued", progress=0, message="Sites KML queued...")
    app_obj = current_app._get_current_object()
    threading.Thread(target=_run_kml_job, args=(app_obj, job_id, "sites", params), daemon=True).start()
    return jsonify({
        "success": True,
        "job_id": job_id,
        "status_url": url_for("doc_bp.kml_job_status", job_id=job_id),
        "download_url": url_for("doc_bp.kml_job_download", job_id=job_id),
    }), 202


@doc_bp.route('/export_kml/sectors/start', methods=['POST'])
@login_required
@csrf_protect
def start_kml_sectors_export():
    job_id = uuid.uuid4().hex
    params = {
        "beam_length_km": request.form.get("beam_length_km", "0.8"),
        "beam_width_deg": request.form.get("beam_width_deg", "40"),
        "beam_color": request.form.get("beam_color", "#0055ff"),
        "region_id": request.form.get("region_id", ""),
        "wilaya_id": request.form.get("wilaya_id", ""),
        "commune_id": request.form.get("commune_id", ""),
        "site_id": request.form.get("site_id", ""),
    }
    _set_kml_job(job_id, status="queued", progress=0, message="Sectors KML queued...")
    app_obj = current_app._get_current_object()
    threading.Thread(target=_run_kml_job, args=(app_obj, job_id, "sectors", params), daemon=True).start()
    return jsonify({
        "success": True,
        "job_id": job_id,
        "status_url": url_for("doc_bp.kml_job_status", job_id=job_id),
        "download_url": url_for("doc_bp.kml_job_download", job_id=job_id),
    }), 202


@doc_bp.route('/export_kml/status/<job_id>', methods=['GET'])
@login_required
def kml_job_status(job_id):
    job = _get_kml_job(job_id)
    if not job:
        return jsonify({"success": False, "message": "KML job not found."}), 404
    return jsonify({
        "success": True,
        "job_id": job_id,
        "status": job.get("status", "unknown"),
        "progress": int(job.get("progress", 0)),
        "message": job.get("message", ""),
        "processed": int(job.get("processed", 0)),
        "total": int(job.get("total", 0)),
        "download_ready": bool(job.get("file_path")),
        "download_url": url_for("doc_bp.kml_job_download", job_id=job_id),
    }), 200


@doc_bp.route('/export_kml/download/<job_id>', methods=['GET'])
@login_required
def kml_job_download(job_id):
    job = _get_kml_job(job_id)
    file_path = job.get("file_path") if job else None
    if not file_path:
        return jsonify({"success": False, "message": "KML file not ready."}), 404
    p = Path(file_path)
    if not p.exists():
        return jsonify({"success": False, "message": "KML file missing."}), 404
    return send_file(
        str(p),
        as_attachment=True,
        download_name=job.get("download_name", p.name),
        mimetype='application/vnd.google-earth.kml+xml',
    )


def _build_sector_placemark_with_options(sector, beam_length_km, beam_width_deg, line_color, poly_color, preloaded_cells=None):
    site = sector.site
    if not site or site.longitude is None or site.latitude is None:
        return ''

    cells = preloaded_cells if preloaded_cells is not None else (
        sector.cells.all() if hasattr(sector.cells, "all") else list(sector.cells)
    )
    antenna_models = sorted({cell.antenna.model for cell in cells if getattr(cell, 'antenna', None) and cell.antenna.model})
    technologies = sorted({str(cell.technology).strip() for cell in cells if getattr(cell, "technology", None)})
    frequencies = sorted({str(cell.frequency).strip() for cell in cells if getattr(cell, "frequency", None)})
    antenna_text = ", ".join(antenna_models) if antenna_models else "N/A"
    technologies_text = " / ".join(technologies) if technologies else "N/A"
    frequencies_text = " / ".join(frequencies) if frequencies else "N/A"
    commune_name = site.commune.name if site.commune else ""

    computed_beamwidth = _sector_beamwidth(sector, cells)
    effective_beamwidth = beam_width_deg if beam_width_deg else computed_beamwidth
    beam_polygon = _sector_beam_polygon(
        sector,
        site,
        beamwidth=effective_beamwidth,
        radius_km=beam_length_km,
        points=20,
    )

    description_html = (
        "<![CDATA["
        "<div style='margin:0;padding:0;font-family:Segoe UI,Arial,sans-serif;font-size:12px;color:#0f172a;min-width:320px;'>"
        "<table style='width:100%;border-collapse:separate;border-spacing:0;border:1px solid #d1d5db;border-radius:10px;overflow:hidden;'>"
        f"<tr style='background:#0f172a;color:#f8fafc;'><td style='padding:8px 10px;font-weight:600;font-size:12px;width:38%;letter-spacing:0.2px;'>SECTOR</td><td style='padding:8px 10px;font-weight:700;font-size:13px;'>{escape(sector.code_sector or '')}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>SITE NAME</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(site.name or '')}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>COMMUNE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(commune_name)}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>LONGITUDE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.longitude or ''))}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>LATITUDE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.latitude or ''))}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>ALTITUDE</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(site.altitude or ''))}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>AZ</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(sector.azimuth or ''))}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>HBA</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(str(sector.hba or ''))}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>ANTENNAS</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(antenna_text)}</td></tr>"
        f"<tr style='background:#f8fafc;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>TECHNOLOGIES</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(technologies_text)}</td></tr>"
        f"<tr style='background:#eef2ff;'><td style='padding:8px 10px;font-weight:600;border-top:1px solid #e5e7eb;'>FREQUENCIES</td><td style='padding:8px 10px;border-top:1px solid #e5e7eb;'>{escape(frequencies_text)}</td></tr>"
        "</table>"
        "</div>"
        "]]>"
    )

    style = (
        "      <Style>\n"
        f"        <LineStyle><color>{line_color}</color><width>1.2</width></LineStyle>\n"
        f"        <PolyStyle><color>{poly_color}</color></PolyStyle>\n"
        "        <BalloonStyle>\n"
        "          <text>$[description]</text>\n"
        "        </BalloonStyle>\n"
        "      </Style>\n"
    )

    return (
        '    <Placemark>\n'
        f'      <name>{escape(sector.code_sector or "Sector")}</name>\n'
        '      <Snippet maxLines="0"></Snippet>\n'
        f'      <description>{description_html}</description>\n'
        f"{style}"
        '      <Polygon>\n'
        '        <outerBoundaryIs><LinearRing><coordinates>\n'
        f'          {beam_polygon}\n'
        '        </coordinates></LinearRing></outerBoundaryIs>\n'
        '      </Polygon>\n'
        '    </Placemark>\n'
    )
